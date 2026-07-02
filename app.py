"""Blockchain.com DeFi Wallet clone — Flask backend.

Routes:
  /                        redirect to /login if not authed
  /login                   login page (German)
  /logout                  logout
  /wallet/home             Startseite
  /wallet/assets           Vermögenswerte
  /wallet/currency/<sym>   TRON detail page (chart)
  /wallet/activity         Aktivität
  /admin/login             admin login
  /admin                   admin dashboard
  /admin/upload            POST xlsx upload
  /api/me                  current user info (JSON)
  /api/wallet              wallet data (JSON)
  /api/activity            activity feed (JSON)

Users are loaded from data/users.xlsx (sheet "users": email, password, name).
"""
import io
import os
import secrets
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for, session,
    jsonify, flash, abort,
)
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
USERS_XLSX = os.path.join(DATA_DIR, "users.xlsx")
SESSIONS_XLSX = os.path.join(DATA_DIR, "sessions.xlsx")
ACTIVITY_XLSX = os.path.join(DATA_DIR, "activity.xlsx")
WALLETS_XLSX = os.path.join(DATA_DIR, "wallets.xlsx")
SETTINGS_XLSX = os.path.join(DATA_DIR, "settings.xlsx")

ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "change-me-now")

# Templates and static live under ./app/
APP_ROOT = os.path.dirname(os.path.abspath(__file__))
app = Flask(
    __name__,
    template_folder=os.path.join(APP_ROOT, "app", "templates"),
    static_folder=os.path.join(APP_ROOT, "app", "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB upload cap


# ----------------------------------------------------------------------------
# Jinja filters — German number formatting
#   1.234,56 €   not   €1,234.56
# ----------------------------------------------------------------------------
import re as _re

def _de_fmt(value, decimals):
    """Format a number with German thousands (period) and decimals (comma)."""
    if value is None or value == "":
        return ""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return str(value)
    if n != n:  # NaN
        return ""
    # Negative → keep sign in front, drop the minus prefix from the formatted part
    sign = "-" if n < 0 else ""
    n = abs(n)
    s = f"{n:,.{decimals}f}"  # uses US format → 1,234.56
    # Swap , and . → 1.234,56
    s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"{sign}{s}"


@app.template_filter("eur")
def _eur(value):
    """EUR amount, German format. 1234567.89 → 1.234.567,89 €"""
    return f"{_de_fmt(value, 2)} €"


@app.template_filter("money")
def _money(value):
    """Same as `eur` but without the € suffix. 1234567.89 → 1.234.567,89"""
    return _de_fmt(value, 2)


@app.template_filter("eur_signed")
def _eur_signed(value):
    """EUR with sign prefix. -1234.56 → −1.234,56 €"""
    if value is None or value == "":
        return ""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return str(value)
    sign = "−" if n < 0 else ("" if n == 0 else "+")
    n = abs(n)
    s = f"{n:,.2f}".replace(",", "§").replace(".", ",").replace("§", ".")
    return f"{sign}{s} €"


@app.template_filter("qty")
def _qty(value, decimals=8):
    """Crypto quantity, German format. 0.025 → 0,02500000"""
    return _de_fmt(value, int(decimals))


@app.template_filter("pct")
def _pct(value, decimals=2):
    """Percentage with sign. 1.84 → +1,84 %"""
    if value is None or value == "":
        return ""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return str(value)
    sign = "−" if n < 0 else ("" if n == 0 else "+")
    n = abs(n)
    s = f"{n:,.{int(decimals)}f}".replace(",", "§").replace(".", ",").replace("§", ".")
    return f"{sign}{s} %"


@app.template_filter("addr")
def _addr(value):
    """Crypto address — ellipsize the middle. 'bc1qabc...xyz' → 'bc1qabc…xyz'"""
    if not value:
        return "—"
    s = str(value)
    if len(s) <= 18:
        return s
    return s[:10] + "…" + s[-6:]

os.makedirs(DATA_DIR, exist_ok=True)


# ----------------------------------------------------------------------------
# Storage layer — xlsx-backed, simple. For ~hundreds of users this is plenty.
# ----------------------------------------------------------------------------
def _read_xlsx(path, sheet):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    return out


def _write_xlsx(path, sheet, rows, header):
    """Write a fresh xlsx with header + rows (list of dicts)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append([r.get(h, "") if r.get(h) is not None else "" for h in header])
    wb.save(path)


# ----------------------------------------------------------------------------
# Users
# ----------------------------------------------------------------------------
def _ensure_users_skeleton():
    """If users.xlsx doesn't exist, create it with an admin row."""
    if not os.path.exists(USERS_XLSX):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(["email", "password_hash", "name", "created_at", "last_login"])
        # seed one row so the file is never empty
        ws.append([
            "demo@blockchain-demo.com",
            generate_password_hash("demo1234"),
            "Demo Nutzer",
            datetime.utcnow().isoformat(timespec="seconds"),
            "",
        ])
        wb.save(USERS_XLSX)


def _all_users():
    _ensure_users_skeleton()
    return _read_xlsx(USERS_XLSX, "users")


def _find_user(email):
    email = (email or "").strip().lower()
    for u in _all_users():
        if str(u.get("email", "")).strip().lower() == email:
            return u
    return None


def _update_user(email, **fields):
    users = _all_users()
    email = email.strip().lower()
    for u in users:
        if str(u.get("email", "")).strip().lower() == email:
            u.update(fields)
            _write_xlsx(
                USERS_XLSX, "users", users,
                ["email", "password_hash", "name", "created_at", "last_login"],
            )
            return
    # user not found -> append
    u = {"email": email, "created_at": datetime.utcnow().isoformat(timespec="seconds")}
    u.update(fields)
    users.append(u)
    _write_xlsx(
        USERS_XLSX, "users", users,
        ["email", "password_hash", "name", "created_at", "last_login"],
    )


# ----------------------------------------------------------------------------
# Wallets  (one balance row per user)
# ----------------------------------------------------------------------------
def _ensure_wallets_skeleton():
    """Create wallets.xlsx with the full schema on first run, AND migrate any
    pre-existing file (no address column, missing BTC row, etc.) so the demo
    user has a real BTC balance + address after upgrading."""
    from openpyxl import load_workbook as _lw
    from openpyxl import Workbook as _wb

    btc_price = next((c["price_eur"] for c in COINS if c["symbol"] == "BTC"), 0)
    seed_demo = [
        ("BTC",   "Bitcoin",  0.025 * btc_price, 0.025, "bc1qxy2kgdygjrsqtzq2n0yrf2493p83kkfjhx0wlh"),
        ("ETH",   "Ethereum", 0.00, 0.0, "0x71C7656EC7ab88b098defB751B7401B5f6d8976F"),
        ("USDT",  "Tether",   0.00, 0.0, "TR7NHqjeKQxGTCi8q8ZY4pL8otSzgjD6Sz"),
        ("BNB",   "BNB",      0.00, 0.0, "bnb1grpf0955t0tlt8eaw9g0w78v5q8v3f5d3wqczp"),
        ("SOL",   "Solana",   0.00, 0.0, "7EYnhQoAGqH7ZbRq8HQq8j4xQ4v5v9NQ7vC2vN3o8X7XJ"),
        ("USDC",  "USD Coin", 0.00, 0.0, "0xA0b86991c6218b36c1d19D4a2e9Eb0cE3606eB48"),
        ("XRP",   "XRP",      0.00, 0.0, "rDsbeomae4FXwgQTJp9Rs64Qg9vDiTCdBv"),
        ("ADA",   "Cardano",  0.00, 0.0, "addr1q9zy2kgdygjrsqtzq2n0yrf2493p83kkfjhx0wlh"),
        ("DOGE",  "Dogecoin", 0.00, 0.0, "DH5yaieqoZN36pDV3xcpbwAY7Sa1YQsv7p"),
        ("TRX",   "TRON",     0.00, 0.0, "TQrZ7d8xNhP9xK2yR5hLkQ3jF8m6bC4wYvE"),
("MATIC", "Polygon",   0.00, 0.0, "0x742d35Cc6634C0532925a3b844Bc9e7595f0bEb0"),
        ("DOT",   "Polkadot",  0.00, 0.0, "1FRMM8d8HdJzk6FpZ7j5vW2pGcZ7Y2qJxC8Rf5oKjHn8U"),
        ("BCH",   "Bitcoin Cash", 0.00, 0.0, "qr5y8q7x9t2k4v6r8p3m5n7s9w2y4a6c8e0g2i4k6o8q0u2w4y6a8c0e2g4i6k8m0o2q4s6u8w0y2a4c6e8g0i2k4m6"),
    ]
    seed_addr = {row[0]: row[4] for row in seed_demo}
    seed_qty  = {row[0]: row[3] for row in seed_demo}   # tuple index 3 = qty
    seed_eur  = {row[0]: row[2] for row in seed_demo}   # tuple index 2 = EUR (qty * price)
    seed_name = {row[0]: row[1] for row in seed_demo}

    full_header = ["email", "symbol", "name", "balance_eur", "balance_qty", "qty_unit", "address"]

    if not os.path.exists(WALLETS_XLSX):
        wb = _wb()
        ws = wb.active
        ws.title = "wallets"
        ws.append(full_header)
        for sym, name, eur, qty, addr in seed_demo:
            # seed_demo tuple is (sym, name, eur, qty, addr)
            ws.append(["demo@blockchain-demo.com", sym, name, eur, qty, sym, addr])
        wb.save(WALLETS_XLSX)
        return

    # Migrate an existing file: ensure header has address, ensure demo user
    # has all 12 coin rows (with BTC carrying a balance + address).
    try:
        wb = _lw(WALLETS_XLSX, data_only=True)
    except Exception:
        return
    if "wallets" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["wallets"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return

    header = [str(h or "").strip() for h in rows[0]]
    needs_address = "address" not in header
    if needs_address:
        header.append("address")
    body = [dict(zip(header, list(r) + [""] * max(0, len(header) - len(r)))) for r in rows[1:]]

# For every user, ensure all 12 coins have rows. For the demo user,
    # seed the demo balance + address the first time, OR backfill the address
    # on existing rows that were created before the address column existed.
    users = sorted({str(r["email"]).strip().lower() for r in body if r.get("email")})
    have_by_key = {(str(r.get("email", "")).strip().lower(), str(r.get("symbol", "")).strip().upper()): r for r in body}
    for email in users:
        is_demo = (email == "demo@blockchain-demo.com")
        for sym in (c["symbol"] for c in COINS):
            key = (email, sym)
            if key in have_by_key:
                # Backfill address if missing (pre-migration file).
                row = have_by_key[key]
                if is_demo:
                    if not (row.get("address") or ""):
                        row["address"] = seed_addr.get(sym, "")
                    # If BTC row exists but has no balance, seed the demo amount.
                    if sym == "BTC":
                        cur_qty = float(row.get("balance_qty") or 0)
                        cur_eur = float(row.get("balance_eur") or 0)
                        # Detect the prior swapped-migration bug:
                        # qty near 0.025 with eur ~2180, OR qty ~2180 with eur near 0.025.
                        if abs(cur_qty - 0.025) < 0.001 and abs(cur_eur - 0.025) < 0.001:
                            pass  # already correct, skip
                        elif abs(cur_qty - 0.025) < 0.001 and cur_eur > 1000:
                            # eur is the qty (~2180), qty is the eur (~0.025) — fix
                            row["balance_qty"], row["balance_eur"] = cur_eur, cur_qty
                        elif cur_qty > 1000 and abs(cur_eur - 0.025) < 0.001:
                            # other side of the swap — fix
                            row["balance_qty"], row["balance_eur"] = cur_eur, cur_qty
                        elif cur_qty == 0 and cur_eur == 0:
                            row["balance_qty"] = seed_qty.get("BTC", 0.025)
                            row["balance_eur"] = seed_eur.get("BTC", 0.025 * btc_price)
                continue
            row = {
                "email": email,
                "symbol": sym,
                "name": seed_name.get(sym, sym),
                "balance_eur": seed_eur.get(sym, 0) if is_demo else 0,
                "balance_qty": seed_qty.get(sym, 0)  if is_demo else 0,
                "qty_unit": sym,
                "address": seed_addr.get(sym, "")      if is_demo else "",
            }
            body.append(row)

    # Make sure every row has all 7 keys
    for r in body:
        for k in full_header:
            r.setdefault(k, "")

    _write_xlsx(WALLETS_XLSX, "wallets", body, full_header)


def _wallets_for(email):
    _ensure_wallets_skeleton()
    _ensure_user_wallet_rows(email)
    email = (email or "").strip().lower()
    out = []
    for w in _read_xlsx(WALLETS_XLSX, "wallets"):
        if str(w.get("email", "")).strip().lower() == email:
            out.append(w)
    # Return in canonical COINS order, then anything else
    by_sym = {str(w.get("symbol", "")).upper(): w for w in out}
    ordered = []
    for c in COINS:
        if c["symbol"] in by_sym:
            ordered.append(by_sym[c["symbol"]])
        else:
            ordered.append({"email": email, "symbol": c["symbol"], "name": c["name"],
                            "balance_eur": 0, "balance_qty": 0, "qty_unit": c["symbol"], "address": ""})
    return ordered


def _ensure_user_wallet_rows(email):
    """Make sure email has a row for every coin in COINS (€0 if missing).
    Also fixes known bugs in the demo BTC row: a previous migration swapped the
    qty / eur columns, so we detect and correct them here on every call."""
    email = (email or "").strip().lower()
    if not email:
        return
    rows = _read_xlsx(WALLETS_XLSX, "wallets")
    have = {str(w.get("symbol", "")).upper()
            for w in rows if str(w.get("email", "")).strip().lower() == email}
    changed = False
    # Repair swapped BTC row on disk for the demo user.
    if email == "demo@blockchain-demo.com":
        for r in rows:
            if (str(r.get("email", "")).strip().lower() == "demo@blockchain-demo.com"
                    and str(r.get("symbol", "")).strip().upper() == "BTC"):
                cur_qty = float(r.get("balance_qty") or 0)
                cur_eur = float(r.get("balance_eur") or 0)
                # Two known-bad shapes from the prior bug:
                #   shape A: qty≈0.025, eur>1000   (eur stored as qty, qty stored as eur)
                #   shape B: qty>1000, eur≈0.025
                if (abs(cur_qty - 0.025) < 0.001 and cur_eur > 1000):
                    r["balance_qty"], r["balance_eur"] = cur_eur, cur_qty
                    changed = True
                elif (cur_qty > 1000 and abs(cur_eur - 0.025) < 0.001):
                    r["balance_qty"], r["balance_eur"] = cur_eur, cur_qty
                    changed = True
                # Backfill address if missing.
                if not (r.get("address") or ""):
                    btc_addr = "bc1qxy2kgdygjrsqtzq2n0yrf2493p83kkfjhx0wlh"
                    r["address"] = btc_addr
                    changed = True
                break
    missing = [c for c in COINS if c["symbol"] not in have]
    if missing:
        changed = True
        for c in missing:
            rows.append({"email": email, "symbol": c["symbol"], "name": c["name"],
                         "balance_eur": 0, "balance_qty": 0, "qty_unit": c["symbol"], "address": ""})
    if changed:
        _write_wallets(rows)


def _write_wallets(rows):
    """Persist the wallets sheet including the optional address column."""
    _write_xlsx(
        WALLETS_XLSX, "wallets", rows,
        ["email", "symbol", "name", "balance_eur", "balance_qty", "qty_unit", "address"],
    )


# ----------------------------------------------------------------------------
# System settings (BCH release fee + future config)
# ----------------------------------------------------------------------------
DEFAULT_SETTINGS = {
    # Required EUR-denominated "Freigabe-Gebühr" the sender must pay in BCH
    # to release a pending transaction. Admin-editable on /admin/settings.
    "bch_release_fee_eur": 380.00,
}


def _ensure_settings_skeleton():
    """Seed settings.xlsx with DEFAULT_SETTINGS on first run."""
    if not os.path.exists(SETTINGS_XLSX):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "settings"
        ws.append(["key", "value"])
        for k, v in DEFAULT_SETTINGS.items():
            ws.append([k, str(v)])
        wb.save(SETTINGS_XLSX)


def _get_setting(key, default=None):
    """Read a single setting by key. Returns `default` if missing/bad type."""
    _ensure_settings_skeleton()
    rows = _read_xlsx(SETTINGS_XLSX, "settings")
    for r in rows:
        if str(r.get("key", "")).strip() == key:
            raw = r.get("value")
            if default is None:
                return raw
            try:
                return type(default)(raw)
            except (TypeError, ValueError):
                return default
    return default


def _set_setting(key, value):
    """Upsert a single setting by key (value coerced to str)."""
    _ensure_settings_skeleton()
    rows = _read_xlsx(SETTINGS_XLSX, "settings")
    for r in rows:
        if str(r.get("key", "")).strip() == key:
            r["value"] = str(value)
            _write_xlsx(SETTINGS_XLSX, "settings", rows, ["key", "value"])
            return
    rows.append({"key": key, "value": str(value)})
    _write_xlsx(SETTINGS_XLSX, "settings", rows, ["key", "value"])


def _set_wallet(email, symbol, **fields):
    """Update one wallet row (email + symbol). Recreates the row if missing.
    Supported fields: balance_qty, balance_eur, address, name."""
    email = (email or "").strip().lower()
    symbol = (symbol or "").strip().upper()
    if not email or not symbol:
        return False
    rows = _read_xlsx(WALLETS_XLSX, "wallets")
    universe = COIN_BY_SYM.get(symbol, {"symbol": symbol, "name": symbol})
    found = False
    for r in rows:
        if (str(r.get("email", "")).strip().lower() == email
                and str(r.get("symbol", "")).strip().upper() == symbol):
            if "balance_qty" in fields and fields["balance_qty"] is not None:
                qty = float(fields["balance_qty"])
                r["balance_qty"] = qty
                # auto-recompute EUR from qty × current price unless caller provided both
                if "balance_eur" not in fields or fields["balance_eur"] in (None, ""):
                    r["balance_eur"] = round(qty * universe.get("price_eur", 0), 2)
            if "balance_eur" in fields and fields["balance_eur"] not in (None, ""):
                r["balance_eur"] = float(fields["balance_eur"])
            if "address" in fields:
                r["address"] = str(fields["address"] or "").strip()
            found = True
            break
    if not found:
        new = {
            "email": email,
            "symbol": symbol,
            "name": universe.get("name", symbol),
            "balance_eur": 0,
            "balance_qty": 0,
            "qty_unit": symbol,
            "address": "",
        }
        if "balance_qty" in fields and fields["balance_qty"] is not None:
            new["balance_qty"] = float(fields["balance_qty"])
            new["balance_eur"] = round(new["balance_qty"] * universe.get("price_eur", 0), 2)
        if "balance_eur" in fields and fields["balance_eur"] not in (None, ""):
            new["balance_eur"] = float(fields["balance_eur"])
        if "address" in fields:
            new["address"] = str(fields["address"] or "").strip()
        rows.append(new)
    _write_wallets(rows)
    return True


def _apply_wallet_upload(email, symbol, qty, address):
    """Create-or-update one wallet row from an admin upload."""
    _set_wallet(email, symbol, balance_qty=qty, address=address)


def _networth(email):
    return round(sum(float(w.get("balance_eur") or 0) for w in _wallets_for(email)), 2)


# ----------------------------------------------------------------------------
# Activity feed
# ----------------------------------------------------------------------------
def _ensure_activity_skeleton():
    if not os.path.exists(ACTIVITY_XLSX):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "activity"
        ws.append(["email", "kind", "label", "amount_eur", "amount_qty",
                    "qty_unit", "ts", "status", "address"])
        ws.append([
            "demo@blockchain-demo.com", "received", "Received TRX",
            0.28, 1.0, "TRX",
            datetime(2026, 6, 20, 14, 23).isoformat(timespec="seconds"),
            "completed", "",
        ])
        wb.save(ACTIVITY_XLSX)
        return

    # Migrate an existing file: add `status` + `address` columns if missing so
    # pending sends (Freigabe-Code flow) can be tracked.
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(ACTIVITY_XLSX, data_only=True)
    except Exception:
        return
    if "activity" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["activity"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return
    header = [str(h or "").strip() for h in rows[0]]
    if "status" not in header:
        header.append("status")
    if "address" not in header:
        header.append("address")
    body = [dict(zip(header, list(r) + [""] * max(0, len(header) - len(r)))) for r in rows[1:]]
    # Backfill default values
    for r in body:
        r.setdefault("status", "completed")
        r.setdefault("address", "")
    _write_xlsx(ACTIVITY_XLSX, "activity", body, header)


def _activity_for(email):
    _ensure_activity_skeleton()
    email = (email or "").strip().lower()
    rows = []
    for r in _read_xlsx(ACTIVITY_XLSX, "activity"):
        if str(r.get("email", "")).strip().lower() == email:
            rows.append(r)
    rows.sort(key=lambda r: str(r.get("ts", "")), reverse=True)
    return rows


def _log_activity(email, *, kind, label, amount_eur=0.0, amount_qty=0.0,
                  qty_unit="", status="completed", address=""):
    """Append one activity row. Migrates the sheet if `status`/`address`
    columns are missing on the legacy schema."""
    _ensure_activity_skeleton()
    rows = _read_xlsx(ACTIVITY_XLSX, "activity")
    header = ["email", "kind", "label", "amount_eur", "amount_qty",
              "qty_unit", "ts", "status", "address"]
    rows.append({
        "email": (email or "").strip().lower(),
        "kind": kind,
        "label": label,
        "amount_eur": round(float(amount_eur or 0), 8),
        "amount_qty": round(float(amount_qty or 0), 8),
        "qty_unit": qty_unit or "",
        "ts": datetime.utcnow().isoformat(timespec="seconds"),
        "status": status,
        "address": address or "",
    })
    _write_xlsx(ACTIVITY_XLSX, "activity", rows, header)


# ----------------------------------------------------------------------------
# Auth decorator
# ----------------------------------------------------------------------------
def login_required(fn):
    @wraps(fn)
    def _w(*a, **kw):
        if not session.get("user_email"):
            return redirect(url_for("login", next=request.path))
        return fn(*a, **kw)
    return _w


def admin_required(fn):
    @wraps(fn)
    def _w(*a, **kw):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login", next=request.path))
        return fn(*a, **kw)
    return _w


# ----------------------------------------------------------------------------
# Currency universe (12 popular coins — used by home/assets/DEX/modal)
# ----------------------------------------------------------------------------
COINS = [
    # symbol, name, color, decimals, price_eur (live), 24h change %, market cap EUR
    {"symbol": "BTC",   "name": "Bitcoin",   "color": "#f7931a", "decimals": 8,  "price_eur":   87234.50, "change_24h_pct":  1.84, "market_cap_eur": "1.72T"},
    {"symbol": "ETH",   "name": "Ethereum",  "color": "#627eea", "decimals": 18, "price_eur":    3208.72, "change_24h_pct":  0.92, "market_cap_eur": "385.6B"},
    {"symbol": "USDT",  "name": "Tether",    "color": "#26a17b", "decimals": 6,  "price_eur":       0.92, "change_24h_pct": -0.02, "market_cap_eur": "112.4B"},
    {"symbol": "BNB",   "name": "BNB",       "color": "#f3ba2f", "decimals": 18, "price_eur":     584.30, "change_24h_pct":  2.45, "market_cap_eur":  "85.7B"},
    {"symbol": "SOL",   "name": "Solana",    "color": "#9945ff", "decimals": 9,  "price_eur":     152.18, "change_24h_pct":  3.61, "market_cap_eur":  "70.8B"},
    {"symbol": "USDC",  "name": "USD Coin",  "color": "#2775ca", "decimals": 6,  "price_eur":       0.92, "change_24h_pct":  0.01, "market_cap_eur":  "32.1B"},
    {"symbol": "XRP",   "name": "XRP",       "color": "#23292f", "decimals": 6,  "price_eur":       0.48, "change_24h_pct": -1.27, "market_cap_eur":  "26.3B"},
    {"symbol": "ADA",   "name": "Cardano",   "color": "#0033ad", "decimals": 6,  "price_eur":       0.36, "change_24h_pct": -0.85, "market_cap_eur":  "12.9B"},
    {"symbol": "DOGE",  "name": "Dogecoin",  "color": "#c2a633", "decimals": 8,  "price_eur":       0.12, "change_24h_pct":  4.21, "market_cap_eur":  "17.4B"},
    {"symbol": "TRX",   "name": "TRON",      "color": "#ff060a", "decimals": 6,  "price_eur":       0.28, "change_24h_pct":  0.12, "market_cap_eur":  "24.1B"},
    {"symbol": "MATIC", "name": "Polygon",   "color": "#8247e5", "decimals": 18, "price_eur":       0.42, "change_24h_pct": -2.18, "market_cap_eur":   "3.9B"},
    {"symbol": "DOT",   "name": "Polkadot",  "color": "#e6007a", "decimals": 10, "price_eur":       5.83, "change_24h_pct":  1.05, "market_cap_eur":   "8.2B"},
    {"symbol": "BCH",   "name": "Bitcoin Cash", "color": "#0ac18e", "decimals": 8, "price_eur": 384.50, "change_24h_pct":  1.52, "market_cap_eur":  "7.6B"},
]
COIN_BY_SYM = {c["symbol"]: c for c in COINS}


# ----------------------------------------------------------------------------
# Aktien page — reuses COINS so DEX and Aktien stay consistent
# ----------------------------------------------------------------------------
STOCKS = [
    {
        "ticker": c["symbol"],
        "name": c["name"],
        "color": c["color"],
        "price_eur": c["price_eur"],
        "change_24h_pct": c["change_24h_pct"],
        "market_cap_eur": c["market_cap_eur"],
    }
    for c in COINS
]


# Deterministic pseudo sparkline for each ticker (24 points over a 0..100 viewbox)
# Direction follows the 24h change so up-trending assets have an upward line.
def _sparkline(ticker, change_pct):
    h = abs(hash("spark::" + ticker)) % 1000
    base = [(h * (i + 7)) % 97 for i in range(24)]
    if change_pct >= 0:
        base = sorted(base)
    else:
        base = sorted(base, reverse=True)
    return " ".join(f"{i*4.35:.1f},{30 - v*0.28:.1f}" for i, v in enumerate(base))


for _s in STOCKS:
    _s["sparkline"] = _sparkline(_s["ticker"], _s["change_24h_pct"])


# ----------------------------------------------------------------------------
# Public routes
# ----------------------------------------------------------------------------
@app.route("/")
def index():
    if not session.get("user_email"):
        return redirect(url_for("login"))
    return redirect(url_for("wallet_home"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip()
        pw = request.form.get("password") or ""

        # two-step screen accepts just email -> password screen
        if "password" not in request.form:
            user = _find_user(email)
            if not user:
                return render_template(
                    "login.html", step="email",
                    error="Ungültige E-Mail oder Wallet-ID",
                    email=email,
                )
            return render_template("login.html", step="password", email=email)

        # password step
        user = _find_user(email)
        if not user or not check_password_hash(str(user.get("password_hash") or ""), pw):
            return render_template(
                "login.html", step="password",
                error="Falsches Passwort",
                email=email,
            )

        session.clear()
        session["user_email"] = str(user["email"]).strip().lower()
        session["user_name"] = str(user.get("name") or "").strip() or session["user_email"]
        _update_user(session["user_email"], last_login=datetime.utcnow().isoformat(timespec="seconds"))

        nxt = request.args.get("next") or url_for("wallet_home")
        return redirect(nxt)

    # GET
    return render_template("login.html", step="email")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ----------------------------------------------------------------------------
# Wallet pages
# ----------------------------------------------------------------------------
def _wallet_ctx(email):
    wallets = _wallets_for(email)
    # Sort by EUR balance desc so the home page shows the user's top holdings first.
    wallets.sort(key=lambda w: float(w.get("balance_eur") or 0), reverse=True)
    return {
        "user_email": email,
        "user_name": session.get("user_name") or email,
        "networth": _networth(email),
        "wallets": wallets,
        "coins": COINS,
        "coin_by_sym": COIN_BY_SYM,
    }


@app.route("/wallet/home")
@login_required
def wallet_home():
    ctx = _wallet_ctx(session["user_email"])
    ctx["activity"] = _activity_for(session["user_email"])[:5]
    return render_template("home.html", **ctx)


@app.route("/wallet/assets")
@login_required
def wallet_assets():
    return render_template("assets.html", **_wallet_ctx(session["user_email"]))


@app.route("/wallet/currency/<sym>")
@login_required
def wallet_currency(sym):
    sym = sym.upper()
    wallets = _wallets_for(session["user_email"])
    coin = next((w for w in wallets if str(w.get("symbol", "")).upper() == sym), None)
    # Pull price + 24h from COINS universe (single source of truth)
    universe = COIN_BY_SYM.get(sym, {"symbol": sym, "name": sym, "price_eur": 0, "change_24h_pct": 0, "color": "#888"})
    if not coin:
        coin = {
            "symbol": sym,
            "name": universe["name"],
            "balance_eur": 0,
            "balance_qty": 0,
            "qty_unit": sym,
        }
    coin["price_eur"] = universe["price_eur"]
    coin["change_24h_pct"] = universe["change_24h_pct"]
    coin["color"] = universe["color"]
    ctx = _wallet_ctx(session["user_email"])
    ctx["coin"] = coin
    return render_template("currency.html", **ctx)


@app.route("/wallet/activity")
@login_required
def wallet_activity():
    return render_template(
        "activity.html",
        **_wallet_ctx(session["user_email"]),
        activity=_activity_for(session["user_email"]),
    )


@app.route("/wallet/dex")
@login_required
def wallet_dex():
    """DEX / swap UI — pure presentation, no live pricing yet."""
    ctx = _wallet_ctx(session["user_email"])
    ctx["coins"] = COINS
    ctx["stocks"] = STOCKS
    return render_template("dex.html", **ctx)


@app.route("/wallet/stocks")
@login_required
def wallet_stocks():
    """Tokenised equities UI — mock data from STOCKS."""
    ctx = _wallet_ctx(session["user_email"])
    ctx["stocks"] = STOCKS
    return render_template("stocks.html", **ctx)


# ----------------------------------------------------------------------------
# JSON APIs (frontend can call these if you wire up live data later)
# ----------------------------------------------------------------------------
@app.route("/api/me")
@login_required
def api_me():
    return jsonify({
        "email": session["user_email"],
        "name": session.get("user_name"),
        "networth": _networth(session["user_email"]),
    })


@app.route("/api/wallet")
@login_required
def api_wallet():
    return jsonify({"networth": _networth(session["user_email"]),
                    "wallets": _wallets_for(session["user_email"])})


@app.route("/api/activity")
@login_required
def api_activity():
    return jsonify({"activity": _activity_for(session["user_email"])})


# ----------------------------------------------------------------------------
# Senden (send) — modal submit, pending tx, BCH release fee confirmation
# ----------------------------------------------------------------------------
# Address where the user must send the BCH release fee to.
# This is the "Freigabe-Code" requirement after they initiate any send.
BCH_RELEASE_ADDRESS = "qrm4k7n9p2s5t8v1w4y6a3c5e7g9i1k3m5o7q9s1u3w5y7"

# Allowed address chars per symbol (basic format check, not chain validation).
def _address_ok(symbol, address):
    a = (address or "").strip()
    if not a or len(a) < 10 or len(a) > 128:
        return False
    if symbol in ("BTC", "BCH"):
        # cashaddr (bch:…) or legacy (1…, 3…) or bech32 (bc1…) — alphanumeric only
        return all(c.isalnum() for c in a)
    if symbol in ("ETH", "USDT", "USDC", "BNB", "MATIC"):
        # 0x + 40 hex
        if not a.lower().startswith("0x") or len(a) != 42:
            return False
        return all(c in "0123456789abcdefABCDEF" for c in a[2:])
    if symbol == "SOL":
        return 32 <= len(a) <= 44 and all(c.isalnum() for c in a)
    if symbol == "XRP":
        return a.startswith("r") and 25 <= len(a) <= 35 and all(c.isalnum() for c in a)
    if symbol == "ADA":
        return a.startswith("addr1") and 50 <= len(a) <= 120
    if symbol == "DOGE":
        return a.startswith("D") and 26 <= len(a) <= 40 and all(c.isalnum() for c in a)
    if symbol == "TRX":
        return a.startswith("T") and 26 <= len(a) <= 40 and all(c.isalnum() for c in a)
    if symbol == "DOT":
        return 46 <= len(a) <= 60 and all(c.isalnum() for c in a)
    return all(c.isalnum() for c in a)


@app.route("/wallet/send", methods=["GET", "POST"])
@login_required
def wallet_send():
    """Render the Senden modal (GET, used to populate the form), or accept a
    submission (POST): log a pending transaction + redirect to the BCH release
    page so the user sees the Freigabe-Code requirement."""
    email = session["user_email"]
    wallets = _wallets_for(email)
    if request.method == "POST":
        symbol = (request.form.get("symbol") or "").strip().upper()
        address = (request.form.get("address") or "").strip()
        try:
            amount_qty = float(request.form.get("amount_qty") or 0)
        except ValueError:
            amount_qty = 0.0

        wallet = next((w for w in wallets if str(w.get("symbol", "")).upper() == symbol), None)
        if not wallet:
            flash(f"Unbekannter Coin: {symbol}", "error")
            return redirect(url_for("wallet_home"))
        if amount_qty <= 0:
            flash("Bitte einen Betrag gr\u00f6\u00dfer als 0 angeben.", "error")
            return redirect(url_for("wallet_home"))
        if not _address_ok(symbol, address):
            flash("Empf\u00e4nger-Adresse hat ein ung\u00fcltiges Format.", "error")
            return redirect(url_for("wallet_home"))

        available = float(wallet.get("balance_qty") or 0)
        if amount_qty > available:
            flash("Unzureichendes Guthaben f\u00fcr diesen Coin.", "error")
            return redirect(url_for("wallet_home"))

        price_eur = float(COIN_BY_SYM.get(symbol, {}).get("price_eur") or 0)
        amount_eur = amount_qty * price_eur

        # Log a pending send activity (status='pending').
        _log_activity(email, kind="send", label=f"Senden {symbol}",
                      amount_eur=amount_eur, amount_qty=amount_qty,
                      qty_unit=symbol, status="pending", address=address)

        # Stash the pending tx in the session so the confirm page can render it
        # without re-asking. Keep it small + serializable.
        session["pending_send"] = {
            "symbol": symbol,
            "name": wallet.get("name") or symbol,
            "address": address,
            "amount_qty": amount_qty,
            "amount_eur": amount_eur,
            "price_eur": price_eur,
            "ts": datetime.utcnow().isoformat(timespec="seconds"),
        }
        return redirect(url_for("wallet_send_confirm"))

    # GET — return a JSON dump of wallets for the modal to populate the picker
    return jsonify({"wallets": wallets})


@app.route("/wallet/send/confirm")
@login_required
def wallet_send_confirm():
    """Show the pending transaction + BCH Freigabe-Code requirement."""
    pending = session.pop("pending_send", None)
    if not pending:
        flash("Keine ausstehende \u00dcberweisung.", "error")
        return redirect(url_for("wallet_home"))

    ctx = _wallet_ctx(session["user_email"])
    ctx.update(_freigabe_ctx(pending))
    return render_template("send_confirm.html", **ctx)


# ----------------------------------------------------------------------------
# Unified trade wall — buy / sell / swap all redirect to a shared BCH-wall page
# ----------------------------------------------------------------------------
BCH_RELEASE_ADDRESS = "qrm4k7n9p2s5t8v1w4y6a3c5e7g9i1k3m5o7q9s1u3w5y7"


def _freigabe_ctx(pending):
    """Build the common context for any BCH Freigabe-Code confirmation page."""
    release_fee_eur = float(_get_setting("bch_release_fee_eur", DEFAULT_SETTINGS["bch_release_fee_eur"]))
    bch = COIN_BY_SYM.get("BCH", {"price_eur": 0, "symbol": "BCH", "name": "Bitcoin Cash", "color": "#0ac18e"})
    release_fee_bch = (release_fee_eur / float(bch["price_eur"])) if bch.get("price_eur") else 0
    return {
        "pending": pending,
        "release_fee_eur": round(release_fee_eur, 2),
        "release_fee_bch": release_fee_bch,
        "release_fee_bch_qty": round(release_fee_bch, 6),
        "release_address": BCH_RELEASE_ADDRESS,
        "bch": bch,
    }


def _log_pending_send(email, *, kind, label, symbol, amount_eur, amount_qty, address=""):
    """Log a pending activity row for any trade action (buy/sell/swap/send)."""
    _log_activity(email, kind=kind, label=label, amount_eur=amount_eur,
                  amount_qty=amount_qty, qty_unit=symbol, status="pending", address=address)


@app.route("/wallet/buy", methods=["GET", "POST"])
@login_required
def wallet_buy():
    """Buy crypto: select asset + amount, log a pending buy, redirect to BCH wall."""
    sym = (request.values.get("sym") or "BTC").upper()
    coin = COIN_BY_SYM.get(sym)
    if not coin:
        flash(f"Unbekannter Coin: {sym}", "error")
        return redirect(url_for("wallet_assets"))

    if request.method == "POST":
        try:
            amount_eur = float((request.form.get("amount_eur") or "0").replace(",", "."))
        except ValueError:
            amount_eur = 0.0
        if amount_eur <= 0:
            flash("Bitte einen Betrag größer als 0 angeben.", "error")
            return redirect(url_for("wallet_buy", sym=sym))

        price = float(coin["price_eur"] or 0)
        amount_qty = (amount_eur / price) if price else 0
        _log_pending_send(session["user_email"], kind="buy", label=f"Kaufen {sym}",
                          symbol=sym, amount_eur=amount_eur, amount_qty=amount_qty)
        session["pending_send"] = {
            "kind": "buy",
            "symbol": sym,
            "name": coin["name"],
            "address": "(Wallet-Gutschrift)",
            "amount_qty": amount_qty,
            "amount_eur": amount_eur,
            "price_eur": price,
            "ts": datetime.utcnow().isoformat(timespec="seconds"),
        }
        return redirect(url_for("wallet_trade_confirm"))

    ctx = _wallet_ctx(session["user_email"])
    ctx["coin"] = coin
    ctx["coins"] = COINS
    return render_template("trade_form.html", **ctx, mode="buy", target_sym=sym)


@app.route("/wallet/sell", methods=["GET", "POST"])
@login_required
def wallet_sell():
    """Sell crypto: pick asset you hold + amount, log a pending sell, redirect to wall."""
    sym = (request.values.get("sym") or "BTC").upper()
    if request.method == "POST":
        wallet = next((w for w in _wallets_for(session["user_email"])
                       if w["symbol"].upper() == sym), None)
        if not wallet:
            flash(f"Sie besitzen kein {sym}.", "error")
            return redirect(url_for("wallet_assets"))
        try:
            amount_qty = float((request.form.get("amount_qty") or "0").replace(",", "."))
        except ValueError:
            amount_qty = 0.0
        if amount_qty <= 0:
            flash("Bitte einen Betrag größer als 0 angeben.", "error")
            return redirect(url_for("wallet_sell", sym=sym))
        available = float(wallet.get("balance_qty") or 0)
        if amount_qty > available:
            flash("Unzureichendes Guthaben für diesen Verkauf.", "error")
            return redirect(url_for("wallet_sell", sym=sym))

        coin = COIN_BY_SYM.get(sym, {"name": sym, "price_eur": 0})
        price = float(coin.get("price_eur") or 0)
        amount_eur = amount_qty * price
        _log_pending_send(session["user_email"], kind="sell", label=f"Verkaufen {sym}",
                          symbol=sym, amount_eur=amount_eur, amount_qty=amount_qty)
        session["pending_send"] = {
            "kind": "sell",
            "symbol": sym,
            "name": coin.get("name", sym),
            "address": "(Auszahlung auf Ihr Bankkonto)",
            "amount_qty": amount_qty,
            "amount_eur": amount_eur,
            "price_eur": price,
            "ts": datetime.utcnow().isoformat(timespec="seconds"),
        }
        return redirect(url_for("wallet_trade_confirm"))

    ctx = _wallet_ctx(session["user_email"])
    ctx["coins"] = COINS
    return render_template("trade_form.html", **ctx, mode="sell", target_sym=sym)


@app.route("/wallet/swap", methods=["GET", "POST"])
@login_required
def wallet_swap():
    """Swap crypto: pick from + to + amount, log a pending swap, redirect to wall."""
    if request.method == "POST":
        from_sym = (request.form.get("from_sym") or "").upper()
        to_sym   = (request.form.get("to_sym") or "").upper()
        try:
            amount_qty = float((request.form.get("amount_qty") or "0").replace(",", "."))
        except ValueError:
            amount_qty = 0.0
        if not from_sym or not to_sym or from_sym == to_sym:
            flash("Bitte Ausgangs- und Ziel-Coin wählen.", "error")
            return redirect(url_for("wallet_swap"))
        if amount_qty <= 0:
            flash("Bitte einen Betrag größer als 0 angeben.", "error")
            return redirect(url_for("wallet_swap"))

        from_coin = COIN_BY_SYM.get(from_sym, {"name": from_sym, "price_eur": 0})
        to_coin   = COIN_BY_SYM.get(to_sym,   {"name": to_sym,   "price_eur": 0})
        from_price = float(from_coin.get("price_eur") or 0)
        to_price   = float(to_coin.get("price_eur")   or 0)
        from_eur   = amount_qty * from_price
        to_qty     = (from_eur / to_price) if to_price else 0
        _log_pending_send(session["user_email"], kind="swap",
                          label=f"Tausch {from_sym} → {to_sym}",
                          symbol=from_sym, amount_eur=from_eur, amount_qty=amount_qty)
        session["pending_send"] = {
            "kind": "swap",
            "symbol": from_sym,
            "name": from_coin.get("name", from_sym),
            "to_symbol": to_sym,
            "to_name": to_coin.get("name", to_sym),
            "address": f"Empfang: {to_qty} {to_sym}",
            "amount_qty": amount_qty,
            "amount_qty_to": to_qty,
            "amount_eur": from_eur,
            "price_eur": from_price,
            "ts": datetime.utcnow().isoformat(timespec="seconds"),
        }
        return redirect(url_for("wallet_trade_confirm"))

    ctx = _wallet_ctx(session["user_email"])
    ctx["coins"] = COINS
    return render_template("trade_form.html", **ctx, mode="swap", target_sym="")


@app.route("/wallet/trade/confirm")
@login_required
def wallet_trade_confirm():
    """Shared BCH Freigabe-Code wall for buy / sell / swap."""
    pending = session.pop("pending_send", None)
    if not pending:
        flash("Keine ausstehende Transaktion.", "error")
        return redirect(url_for("wallet_home"))
    ctx = _wallet_ctx(session["user_email"])
    ctx.update(_freigabe_ctx(pending))
    return render_template("send_confirm.html", **ctx)


# ----------------------------------------------------------------------------
# Admin
# ----------------------------------------------------------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = request.form.get("password") or ""
        if u == ADMIN_USER and p == ADMIN_PASS:
            session.clear()
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        return render_template("admin_login.html", error="Falsche Zugangsdaten")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    """Landing page: stats + recent activity + quick links to the user manager."""
    users = _all_users()
    rows = _read_xlsx(WALLETS_XLSX, "wallets")
    rows_act = _read_xlsx(ACTIVITY_XLSX, "activity")

    # Stats
    total_users = len(users)
    today_iso = date.today().isoformat()
    new_today = sum(1 for u in users if str(u.get("created_at") or "")[:10] == today_iso)
    active_7d = sum(
        1 for u in users
        if (u.get("last_login") and str(u["last_login"])[:10] >= (date.today() - __import__("datetime").timedelta(days=7)).isoformat())
    )
    total_wallets = len(rows)
    total_activity = len(rows_act)

    # Most recent 10 users (by created_at desc)
    users_sorted = sorted(users, key=lambda u: str(u.get("created_at") or ""), reverse=True)
    recent = users_sorted[:10]

    return render_template(
        "admin.html",
        active_view="dashboard",
        total_users=total_users,
        new_today=new_today,
        active_7d=active_7d,
        total_wallets=total_wallets,
        total_activity=total_activity,
        recent=recent,
        coins=COINS,
    )


# ----------------------------------------------------------------------------
# Admin — user list (search, filter, sort, pagination, bulk actions)
# ----------------------------------------------------------------------------
def _paginate(items, page, per_page):
    """Slice a list into a page. Returns (page_items, total, total_pages, page, per_page)."""
    try:
        page = max(1, int(page or 1))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = max(5, min(200, int(per_page or 25)))
    except (TypeError, ValueError):
        per_page = 25
    total = len(items)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    start = (page - 1) * per_page
    return items[start:start + per_page], total, total_pages, page, per_page


def _user_query(users, btc_qty_by_email=None, q="", has_btc="all", sort="newest"):
    """Apply search + filter + sort to the user list."""
    q = (q or "").strip().lower()
    btc_qty_by_email = btc_qty_by_email or {}
    out = list(users)

    if q:
        out = [
            u for u in out
            if q in str(u.get("email", "")).lower()
            or q in str(u.get("name", "")).lower()
            or q in str(u.get("last_login", "")).lower()
            or q in str(u.get("created_at", "")).lower()
        ]

    if has_btc == "btc-only":
        out = [u for u in out if btc_qty_by_email.get(str(u.get("email", "")).strip().lower(), 0) > 0]
    elif has_btc == "no-btc":
        out = [u for u in out if btc_qty_by_email.get(str(u.get("email", "")).strip().lower(), 0) == 0]

    if sort == "newest":
        out.sort(key=lambda u: str(u.get("created_at") or ""), reverse=True)
    elif sort == "oldest":
        out.sort(key=lambda u: str(u.get("created_at") or ""))
    elif sort == "name":
        out.sort(key=lambda u: str(u.get("name") or u.get("email") or "").lower())
    elif sort == "email":
        out.sort(key=lambda u: str(u.get("email") or "").lower())
    elif sort == "last-login":
        out.sort(key=lambda u: str(u.get("last_login") or ""), reverse=True)
    return out


@app.route("/admin/users")
@admin_required
def admin_users():
    users = _all_users()
    # precompute BTC qty per user for the "has BTC" filter
    btc_qty_by_email = {}
    for w in _read_xlsx(WALLETS_XLSX, "wallets"):
        if str(w.get("symbol", "")).upper() == "BTC":
            btc_qty_by_email[str(w.get("email", "")).strip().lower()] = float(w.get("balance_qty") or 0)

    q         = request.args.get("q", "")
    has_btc   = request.args.get("btc", "all")
    sort      = request.args.get("sort", "newest")
    page      = request.args.get("page", 1)
    per_page  = request.args.get("per_page", 25)

    filtered = _user_query(users, btc_qty_by_email=btc_qty_by_email, q=q, has_btc=has_btc, sort=sort)
    page_items, total, total_pages, page, per_page = _paginate(filtered, page, per_page)

    # Networth + BTC summary per user (only for the page rows to keep it cheap)
    btc_eur = next((c["price_eur"] for c in COINS if c["symbol"] == "BTC"), 0)
    rows_meta = []
    for u in page_items:
        email = str(u.get("email", "")).strip().lower()
        btc_qty = btc_qty_by_email.get(email, 0)
        rows_meta.append({
            "btc_qty": btc_qty,
            "btc_eur": round(btc_qty * btc_eur, 2),
        })

    return render_template(
        "admin_users.html",
        active_view="users",
        users=page_items,
        rows_meta=rows_meta,
        q=q,
        btc=has_btc,
        sort=sort,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        coins=COINS,
        query_args=request.args,
    )


@app.route("/admin/user/<email>")
@admin_required
def admin_user_detail(email):
    """Detail page: profile, all wallet rows (editable), recent activity, delete."""
    email = (email or "").strip().lower()
    user = _find_user(email)
    if not user:
        flash(f"Kein Benutzer mit E-Mail '{email}'.", "error")
        return redirect(url_for("admin_users"))
    wallets = _wallets_for(email)
    btc_eur = next((c["price_eur"] for c in COINS if c["symbol"] == "BTC"), 0)
    networth_eur = sum(
        float(w.get("balance_qty") or 0) * next((c["price_eur"] for c in COINS if c["symbol"] == str(w.get("symbol", "")).upper()), 0)
        for w in wallets
    )
    activity = _activity_for(email)[:20]
    return render_template(
        "admin_user_detail.html",
        active_view="users",
        user=user,
        wallets=wallets,
        coins=COINS,
        btc_eur=btc_eur,
        networth_eur=round(networth_eur, 2),
        activity=activity,
    )


@app.route("/admin/user/<email>/delete", methods=["POST"])
@admin_required
def admin_user_delete(email):
    email = (email or "").strip().lower()
    rows = _read_xlsx(USERS_XLSX, "users")
    new = [r for r in rows if str(r.get("email", "")).strip().lower() != email]
    if len(new) == len(rows):
        flash(f"Benutzer '{email}' nicht gefunden.", "error")
        return redirect(url_for("admin_users"))
    _write_xlsx(USERS_XLSX, "users", new, ["email", "password_hash", "name", "created_at", "last_login"])
    # also delete their wallets + activity
    for path, sheet in [(WALLETS_XLSX, "wallets"), (ACTIVITY_XLSX, "activity")]:
        ws_rows = _read_xlsx(path, sheet)
        kept = [r for r in ws_rows if str(r.get("email", "")).strip().lower() != email]
        # keep header
        if sheet == "wallets":
            _write_wallets(kept)
        else:
            _write_xlsx(path, sheet, kept,
                        ["email", "kind", "label", "amount_eur", "amount_qty", "qty_unit", "ts"])
    flash(f"Benutzer '{email}' gelöscht.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/user/<email>/activity/delete/<ts>", methods=["POST"])
@admin_required
def admin_activity_delete(email, ts):
    """Delete a single activity row identified by (email, ts).

    ts comes from the URL — we accept any string the row was keyed on.
    """
    email = (email or "").strip().lower()
    target_ts = (ts or "").strip()
    rows = _read_xlsx(ACTIVITY_XLSX, "activity")
    kept = [
        r for r in rows
        if not (str(r.get("email", "")).strip().lower() == email
                and str(r.get("ts", "")).strip() == target_ts)
    ]
    if len(kept) == len(rows):
        flash("Aktivität-Eintrag nicht gefunden.", "error")
    else:
        _write_xlsx(ACTIVITY_XLSX, "activity", kept, [
            "email", "kind", "label", "amount_eur", "amount_qty", "qty_unit", "ts", "status", "address"
        ])
        flash("Aktivität-Eintrag gelöscht.", "success")
    return redirect(url_for("admin_user_detail", email=email) + "#activity")


@app.route("/admin/user/<email>/activity/edit", methods=["POST"])
@admin_required
def admin_activity_edit(email):
    """Edit a single activity row in place: label, amount_eur, amount_qty, status."""
    email = (email or "").strip().lower()
    target_ts = (request.form.get("ts") or "").strip()
    rows = _read_xlsx(ACTIVITY_XLSX, "activity")
    found = False
    for r in rows:
        if (str(r.get("email", "")).strip().lower() == email
                and str(r.get("ts", "")).strip() == target_ts):
            r["label"] = (request.form.get("label") or r.get("label", "")).strip()
            try:
                r["amount_eur"] = round(float((request.form.get("amount_eur") or "0").replace(",", ".")), 8)
            except ValueError:
                r["amount_eur"] = r.get("amount_eur", 0)
            try:
                r["amount_qty"] = round(float((request.form.get("amount_qty") or "0").replace(",", ".")), 8)
            except ValueError:
                r["amount_qty"] = r.get("amount_qty", 0)
            r["status"] = (request.form.get("status") or r.get("status") or "completed").strip()
            found = True
            break
    if not found:
        flash("Aktivität-Eintrag nicht gefunden.", "error")
    else:
        _write_xlsx(ACTIVITY_XLSX, "activity", rows, [
            "email", "kind", "label", "amount_eur", "amount_qty", "qty_unit", "ts", "status", "address"
        ])
        flash("Aktivität-Eintrag aktualisiert.", "success")
    return redirect(url_for("admin_user_detail", email=email) + "#activity")


@app.route("/admin/users/bulk", methods=["POST"])
@admin_required
def admin_users_bulk():
    """Bulk actions: delete | reset_password | export."""
    action = request.form.get("action", "")
    emails = [e.strip().lower() for e in request.form.getlist("emails") if e.strip()]
    if not emails:
        flash("Keine Benutzer ausgewählt.", "error")
        return redirect(request.referrer or url_for("admin_users"))

    if action == "delete":
        users_rows = _read_xlsx(USERS_XLSX, "users")
        new_users = [r for r in users_rows if str(r.get("email", "")).strip().lower() not in emails]
        _write_xlsx(USERS_XLSX, "users", new_users, ["email", "password_hash", "name", "created_at", "last_login"])
        # cascade-delete wallets + activity for the deleted emails
        for path, sheet in [(WALLETS_XLSX, "wallets"), (ACTIVITY_XLSX, "activity")]:
            kept = [r for r in _read_xlsx(path, sheet) if str(r.get("email", "")).strip().lower() not in emails]
            if sheet == "wallets":
                _write_wallets(kept)
            else:
                _write_xlsx(path, sheet, kept,
                            ["email", "kind", "label", "amount_eur", "amount_qty", "qty_unit", "ts"])
        flash(f"{len(emails)} Benutzer gelöscht.", "success")
        return redirect(url_for("admin_users"))

    if action == "reset_password":
        from secrets import token_urlsafe
        new_pw = token_urlsafe(8)
        pw_hash = generate_password_hash(new_pw)
        rows = _read_xlsx(USERS_XLSX, "users")
        n = 0
        for r in rows:
            if str(r.get("email", "")).strip().lower() in emails:
                r["password_hash"] = pw_hash
                n += 1
        _write_xlsx(USERS_XLSX, "users", rows, ["email", "password_hash", "name", "created_at", "last_login"])
        # surface the generated password once (won't be retrievable later)
        flash(f"Neues Passwort für {n} Benutzer: <code>{new_pw}</code>", "success")
        return redirect(url_for("admin_users"))

    if action == "export":
        # stream a CSV of the selected users + their BTC balance
        import csv, io as _io
        btc_eur = next((c["price_eur"] for c in COINS if c["symbol"] == "BTC"), 0)
        btc_by_email = {}
        for w in _read_xlsx(WALLETS_XLSX, "wallets"):
            if str(w.get("symbol", "")).upper() == "BTC":
                btc_by_email[str(w.get("email", "")).strip().lower()] = float(w.get("balance_qty") or 0)
        users_rows = _read_xlsx(USERS_XLSX, "users")
        selected = [r for r in users_rows if str(r.get("email", "")).strip().lower() in set(emails)]
        buf = _io.StringIO()
        w = csv.writer(buf)
        w.writerow(["email", "name", "created_at", "last_login", "btc_qty", "btc_eur"])
        for r in selected:
            email = str(r.get("email", "")).strip().lower()
            btc_q = btc_by_email.get(email, 0)
            w.writerow([email, r.get("name", ""), str(r.get("created_at", ""))[:10],
                        str(r.get("last_login", ""))[:10], btc_q, round(btc_q * btc_eur, 2)])
        from flask import Response
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": 'attachment; filename="users_export.csv"'},
        )

    flash(f"Unbekannte Aktion: {action}", "error")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/export.csv")
@admin_required
def admin_users_export_all():
    """Export the currently-filtered user list to CSV (reuses admin_users query)."""
    users = _all_users()
    btc_qty_by_email = {}
    for w in _read_xlsx(WALLETS_XLSX, "wallets"):
        if str(w.get("symbol", "")).upper() == "BTC":
            btc_qty_by_email[str(w.get("email", "")).strip().lower()] = float(w.get("balance_qty") or 0)

    q         = request.args.get("q", "")
    has_btc   = request.args.get("btc", "all")
    sort      = request.args.get("sort", "newest")
    filtered = _user_query(users, btc_qty_by_email=btc_qty_by_email, q=q, has_btc=has_btc, sort=sort)

    import csv, io as _io
    btc_eur = next((c["price_eur"] for c in COINS if c["symbol"] == "BTC"), 0)
    buf = _io.StringIO()
    cw = csv.writer(buf)
    cw.writerow(["email", "name", "created_at", "last_login", "btc_qty", "btc_eur"])
    for r in filtered:
        email = str(r.get("email", "")).strip().lower()
        btc_q = btc_qty_by_email.get(email, 0)
        cw.writerow([email, r.get("name", ""), str(r.get("created_at", ""))[:10],
                     str(r.get("last_login", ""))[:10], btc_q, round(btc_q * btc_eur, 2)])
    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="users_export.csv"'},
    )


@app.route("/admin/wallet", methods=["POST"])
@admin_required
def admin_wallet_edit():
    """Inline edit of a single (email, symbol) wallet row."""
    email = (request.form.get("email") or "").strip().lower()
    symbol = (request.form.get("symbol") or "").strip().upper()
    qty_raw = request.form.get("balance_qty", "").strip()
    address = request.form.get("address", "").strip()

    if not email or not symbol:
        flash("E-Mail und Symbol sind erforderlich.", "error")
        return redirect(url_for("admin_dashboard"))

    try:
        qty = float(qty_raw) if qty_raw else 0.0
    except ValueError:
        flash(f"Ungültiger Betrag für {email} / {symbol}: {qty_raw!r}", "error")
        return redirect(url_for("admin_dashboard"))

    _set_wallet(email, symbol, balance_qty=qty, address=address)
    flash(f"{symbol}-Guthaben für {email} aktualisiert: {qty} {symbol}", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/settings", methods=["GET", "POST"])
@admin_required
def admin_settings():
    """Edit the BCH release fee + view all system settings."""
    _ensure_settings_skeleton()
    if request.method == "POST":
        # BCH release fee (EUR)
        try:
            fee = float((request.form.get("bch_release_fee_eur") or "0").replace(",", "."))
            if fee < 0:
                raise ValueError
        except ValueError:
            flash("Ungültiger Wert für die BCH-Freigabegebühr.", "error")
            return redirect(url_for("admin_settings"))
        _set_setting("bch_release_fee_eur", round(fee, 2))

        # Save arbitrary extra settings if posted (key/value pairs)
        for k, v in request.form.items():
            if k.startswith("extra_"):
                real_key = k[len("extra_"):]
                if real_key:
                    _set_setting(real_key, v.strip())

        flash(f"BCH-Freigabegebühr auf {round(fee, 2)} € aktualisiert.", "success")
        return redirect(url_for("admin_settings"))

    # GET — render settings page
    settings = {}
    for r in _read_xlsx(SETTINGS_XLSX, "settings"):
        settings[str(r.get("key", "")).strip()] = r.get("value")
    return render_template(
        "admin_settings.html",
        active_view="settings",
        settings=settings,
        bch_release_fee_eur=settings.get("bch_release_fee_eur", DEFAULT_SETTINGS["bch_release_fee_eur"]),
        bch_price_eur=COIN_BY_SYM.get("BCH", {}).get("price_eur", 0),
        coins=COINS,
    )


@app.route("/admin/upload", methods=["POST"])
@admin_required
def admin_upload():
    """Accept xlsx with one of two formats:

    Format A (legacy user import):
      email, password, [name]

    Format B (wallet import — first_name, last_name, optional email,
      optional password, plus balance + address columns per coin):
      first_name, last_name, [email], [password],
      btc_value, btc_address, eth_value, eth_address, usdt_value, usdt_address, ...

    Coin columns are detected automatically: any column whose name ends in
    "_value" is treated as that coin's balance, and the matching
    "<coin>_address" column as that coin's wallet address.
    """
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Keine Datei hochgeladen.", "error")
        return redirect(url_for("admin_dashboard"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("Nur .xlsx-Dateien werden akzeptiert.", "error")
        return redirect(url_for("admin_dashboard"))

    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        flash(f"Datei konnte nicht gelesen werden: {e}", "error")
        return redirect(url_for("admin_dashboard"))

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash("Leere Datei.", "error")
        return redirect(url_for("admin_dashboard"))

    header = [str(h or "").strip().lower() for h in rows[0]]
    header_set = set(header)

    # Detect wallet-import format
    has_wallet_cols = any(c.endswith("_value") for c in header)
    has_first_name = "first_name" in header_set
    has_last_name  = "last_name" in header_set

    if has_wallet_cols or (has_first_name and has_last_name):
        return _import_wallet_xlsx(header, rows)
    return _import_legacy_user_xlsx(header, rows)


def _import_legacy_user_xlsx(header, rows):
    """Legacy format: email, password, name (optional)."""
    required = {"email", "password"}
    missing = required - set(header)
    if missing:
        flash(f"Spalten fehlen: {', '.join(sorted(missing))}. Erforderlich: email, password, (name optional).", "error")
        return redirect(url_for("admin_dashboard"))

    idx = {h: i for i, h in enumerate(header)}
    added = updated = skipped = errors = 0
    error_lines = []
    for lineno, r in enumerate(rows[1:], start=2):
        if r is None or all(c is None for c in r):
            continue
        email = str(r[idx["email"]] or "").strip().lower()
        password = r[idx["password"]]
        name = str(r[idx["name"]] or "").strip() if "name" in idx else ""

        if not email or not password:
            skipped += 1
            continue

        if not email.count("@") or "." not in email.split("@")[-1]:
            errors += 1
            error_lines.append(f"Zeile {lineno}: ungültige E-Mail '{email}'")
            continue

        pw_hash = generate_password_hash(str(password))
        existing = _find_user(email)
        if existing:
            _update_user(email, password_hash=pw_hash, name=name or existing.get("name", ""))
            updated += 1
        else:
            _update_user(
                email, password_hash=pw_hash, name=name or email.split("@")[0],
                created_at=datetime.utcnow().isoformat(timespec="seconds"),
                last_login="",
            )
            added += 1

    msg = f"Import fertig: {added} hinzugefügt, {updated} aktualisiert, {skipped} übersprungen."
    if errors:
        msg += f" {errors} Fehler: " + "; ".join(error_lines[:5])
        if len(error_lines) > 5:
            msg += f" (+{len(error_lines) - 5} weitere)"
        flash(msg, "error")
    else:
        flash(msg, "success")
    return redirect(url_for("admin_dashboard"))


def _import_wallet_xlsx(header, rows):
    """Format B: first_name, last_name, [email], [password], <sym>_value, <sym>_address."""
    idx = {h: i for i, h in enumerate(header)}

    # Discover coin columns
    value_cols  = [h for h in header if h.endswith("_value")]
    address_cols = {h[:-len("_value")] + "_address": None for h in value_cols}
    coins_in_file = []
    for col in value_cols:
        sym = col[:-len("_value")].upper()
        if sym in COIN_BY_SYM:
            coins_in_file.append(sym)
    if not coins_in_file:
        flash("Wallet-Import: keine bekannten <coin>_value Spalten gefunden.", "error")
        return redirect(url_for("admin_dashboard"))

    added = updated = wallets_set = errors = skipped = 0
    error_lines = []

    for lineno, r in enumerate(rows[1:], start=2):
        if r is None or all(c is None for c in r):
            continue

        # Email: prefer explicit column, else build from first_name.last_name@example.com
        email_raw = ""
        if "email" in idx:
            email_raw = str(r[idx["email"]] or "").strip().lower()
        first = str(r[idx["first_name"]] or "").strip() if "first_name" in idx else ""
        last  = str(r[idx["last_name"]]  or "").strip() if "last_name"  in idx else ""
        full_name = (first + " " + last).strip()
        if not email_raw:
            if first and last:
                email_raw = (first + "." + last).lower().replace(" ", ".") + "@example.com"
            else:
                error_lines.append(f"Zeile {lineno}: keine E-Mail und kein Vor-/Nachname")
                errors += 1
                continue
        # Build the email + add user if missing
        password = "demo1234"
        if "password" in idx and r[idx["password"]]:
            password = str(r[idx["password"]])
        pw_hash = generate_password_hash(password)
        if _find_user(email_raw):
            _update_user(email_raw, name=full_name)
            updated += 1
        else:
            _update_user(
                email_raw,
                password_hash=pw_hash,
                name=full_name or email_raw.split("@")[0],
                created_at=datetime.utcnow().isoformat(timespec="seconds"),
                last_login="",
            )
            added += 1

        # Apply each coin row
        for sym in coins_in_file:
            v_col = sym.lower() + "_value"
            a_col = sym.lower() + "_address"
            v_raw = r[idx[v_col]] if v_col in idx else None
            a_raw = r[idx[a_col]] if a_col in idx else None
            try:
                qty = float(v_raw) if v_raw not in (None, "") else 0.0
            except (TypeError, ValueError):
                qty = 0.0
            address = str(a_raw or "").strip()
            _set_wallet(email_raw, sym, balance_qty=qty, address=address)
            wallets_set += 1

    msg = (f"Wallet-Import: {added} Nutzer neu, {updated} aktualisiert, "
           f"{wallets_set} Wallet-Zeilen gesetzt.")
    if errors:
        msg += f" {errors} Fehler: " + "; ".join(error_lines[:5])
        if len(error_lines) > 5:
            msg += f" (+{len(error_lines) - 5} weitere)"
        flash(msg, "error")
    else:
        flash(msg, "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/sample")
@admin_required
def admin_sample():
    """Download a sample xlsx that demos BOTH upload formats."""
    from openpyxl import Workbook
    wb = Workbook()

    # Sheet 1: legacy user import
    ws1 = wb.active
    ws1.title = "users"
    ws1.append(["email", "password", "name"])
    ws1.append(["max@example.com", "geheim123", "Max Mustermann"])
    ws1.append(["anna@example.com", "nocheins", "Anna Schmidt"])

    # Sheet 2: wallet import (first_name / last_name / coin balances + addresses)
    ws2 = wb.create_sheet("wallets")
    header = ["first_name", "last_name", "email", "password"]
    for c in COINS:
        header.append(f"{c['symbol'].lower()}_value")
        header.append(f"{c['symbol'].lower()}_address")
    ws2.append(header)
    ws2.append([
        "Max", "Mustermann", "max@example.com", "demo1234",
        0.5, "bc1qexampleaddress0000000000000000000000000000000",
        4.0, "0x1111111111111111111111111111111111111111",
        0, "",
        0, "",
    ])
    ws2.append([
        "Anna", "Schmidt", "anna@example.com", "demo1234",
        0.1, "bc1qexampleaddress0000000000000000000000000000000",
        0, "",
        100, "TR7NHqjeKQxGTCi8q8ZY4pL8otSzgjD6Sz",
        0, "",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name="blockwall_sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------------------------------
# Bootstrap
# ----------------------------------------------------------------------------
_ensure_users_skeleton()
_ensure_wallets_skeleton()
_ensure_activity_skeleton()


@app.context_processor
def _inject_globals():
    """Make COINS available to every template (used by modal, sidebar, etc.)."""
    return {
        "COINS": COINS,
        "COIN_BY_SYM": COIN_BY_SYM,
        "STOCKS": STOCKS,
    }


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)